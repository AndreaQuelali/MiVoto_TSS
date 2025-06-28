import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import sys
import os

# Agregar el directorio src al path para importar el modelo
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from models.electoral_model import ModeloPredictivoElectoral
from utils.electoral_utils import verificar_segunda_vuelta, calcular_escanos, simular_segunda_vuelta
from config.settings import DATOS_HISTORICOS_DEFAULT, ENCUESTAS_2025_DEFAULT

class ExcelElectoralModel:
    def __init__(self):
        # Datos iniciales (corregidos)
        self.datos_historicos = DATOS_HISTORICOS_DEFAULT
        self.encuestas_2025 = ENCUESTAS_2025_DEFAULT
        
        # Configuración del modelo
        self.peso_historico = 0.4
        self.peso_encuestas = 0.6
        self.margen_error = 0.03
        self.tendencia_ajuste = "Conservar"
        self.umbral_minimo = 0.03
        self.total_senadores = 36
        self.total_diputados = 130
        
        # Crear el modelo predictivo
        self.modelo = ModeloPredictivoElectoral()
        self.modelo.cargar_datos_historicos(self.datos_historicos)
        self.modelo.cargar_encuestas(self.encuestas_2025)
        self.modelo.configurar_parametros(
            self.peso_historico, 
            self.peso_encuestas, 
            self.margen_error, 
            self.tendencia_ajuste, 
            self.umbral_minimo
        )
        
        # Ejecutar predicción
        self.modelo.ejecutar_prediccion()
        self.resultados = self.modelo.obtener_resultados()
        
        # Crear el libro de Excel
        self.wb = Workbook()
        self.ws_datos = self.wb.active
        self.ws_datos.title = "Datos y Configuración"
        
        # Crear otras hojas
        self.ws_prediccion = self.wb.create_sheet("Predicción 2025")
        self.ws_escanos = self.wb.create_sheet("Distribución Escaños")
        self.ws_segunda_vuelta = self.wb.create_sheet("Segunda Vuelta")
        
        # Estilos
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        self.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        self.center_alignment = Alignment(horizontal='center')
    
    def aplicar_estilo_celda(self, cell, is_header=False, is_subheader=False):
        """Aplica estilos a una celda"""
        cell.border = self.border
        if is_header:
            cell.fill = self.header_fill
            cell.font = self.header_font
        elif is_subheader:
            cell.fill = self.subheader_fill
        cell.alignment = self.center_alignment
    
    def crear_hoja_datos(self):
        """Crea la hoja con datos históricos y encuestas"""
        ws = self.ws_datos
        
        # Unificar todos los partidos de históricos y encuestas
        all_parties = sorted(
            set(
                p for data in self.datos_historicos.values() for p in data.keys()
            ).union(
                p for data in self.encuestas_2025.values() for p in data.keys()
            )
        )
        
        # Datos históricos
        ws['A1'] = "Datos Históricos de Elecciones en Bolivia"
        ws.merge_cells('A1:F1')
        self.aplicar_estilo_celda(ws['A1'], is_header=True)
        
        # Encabezados
        ws['A2'] = "Año"
        for col, party in enumerate(all_parties, start=1):
            ws.cell(row=2, column=col+1, value=party)
            self.aplicar_estilo_celda(ws.cell(row=2, column=col+1), is_header=True)
        
        # Datos históricos
        for row, (year, data) in enumerate(sorted(self.datos_historicos.items()), start=3):
            ws.cell(row=row, column=1, value=year)
            for col, party in enumerate(all_parties, start=1):
                ws.cell(row=row, column=col+1, value=data.get(party, 0))
                self.aplicar_estilo_celda(ws.cell(row=row, column=col+1))
        
        # Encuestas 2025
        start_row = len(self.datos_historicos) + 5
        ws.cell(row=start_row, column=1, value="Encuestas de Intención de Voto 2025")
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.aplicar_estilo_celda(ws.cell(row=start_row, column=1), is_header=True)
        
        # Encabezados encuestas
        ws.cell(row=start_row+1, column=1, value="Encuesta")
        for col, party in enumerate(all_parties, start=1):
            ws.cell(row=start_row+1, column=col+1, value=party)
            self.aplicar_estilo_celda(ws.cell(row=start_row+1, column=col+1), is_header=True)
        
        # Datos encuestas
        for row, (survey, data) in enumerate(self.encuestas_2025.items(), start=start_row+2):
            ws.cell(row=row, column=1, value=survey)
            for col, party in enumerate(all_parties, start=1):
                ws.cell(row=row, column=col+1, value=data.get(party, 0))
                self.aplicar_estilo_celda(ws.cell(row=row, column=col+1))
        
        # Configuración del modelo
        config_start_row = start_row + len(self.encuestas_2025) + 3
        ws.cell(row=config_start_row, column=1, value="Configuración del Modelo Predictivo")
        ws.merge_cells(f'A{config_start_row}:B{config_start_row}')
        self.aplicar_estilo_celda(ws.cell(row=config_start_row, column=1), is_header=True)
        
        config_data = [
            ("Peso datos históricos", self.peso_historico),
            ("Peso encuestas 2025", self.peso_encuestas),
            ("Margen de error", self.margen_error),
            ("Umbral mínimo para escaños", self.umbral_minimo),
            ("Tendencia de ajuste", self.tendencia_ajuste),
            ("Total Senadores", self.total_senadores),
            ("Total Diputados", self.total_diputados)
        ]
        
        for row, (label, value) in enumerate(config_data, start=config_start_row+1):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            self.aplicar_estilo_celda(ws.cell(row=row, column=1), is_subheader=True)
            self.aplicar_estilo_celda(ws.cell(row=row, column=2))
        
        # Ajustar anchos de columna
        for col_idx in range(1, len(all_parties) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
    
    def get_config_row(self, config_name):
        """Obtiene la fila donde está un parámetro de configuración"""
        ws = self.ws_datos
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value == config_name:
                return row
        return 0
    
    def crear_hoja_prediccion(self):
        """Crea la hoja con la predicción de votos y fórmulas"""
        ws = self.ws_prediccion
        
        # Título
        ws['A1'] = "Predicción Electoral Bolivia 2025"
        ws.merge_cells('A1:D1')
        self.aplicar_estilo_celda(ws['A1'], is_header=True)
        
        # Datos históricos recientes (última elección)
        ultimo_anio = max(self.datos_historicos.keys())
        datos_recientes = self.datos_historicos[ultimo_anio]
        
        # Promedio de encuestas
        all_parties = set(datos_recientes.keys()).union(*[e.keys() for e in self.encuestas_2025.values()])
        promedios_encuestas = {}
        
        for p in all_parties:
            valores = [e.get(p, 0) for e in self.encuestas_2025.values()]
            promedios_encuestas[p] = np.mean(valores) if valores else 0
        
        # Encabezados
        headers = ["Partido", "Datos Históricos", "Promedio Encuestas", "Predicción 2025"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
            self.aplicar_estilo_celda(ws.cell(row=2, column=col), is_header=True)
        
        # Fórmulas y datos simplificadas
        for row, party in enumerate(all_parties, start=3):
            ws.cell(row=row, column=1, value=party)
            ws.cell(row=row, column=2, value=datos_recientes.get(party, 0))
            ws.cell(row=row, column=3, value=promedios_encuestas.get(party, 0))
            
            # Fórmula de predicción simplificada
            peso_hist = 0.4  # Valor fijo para evitar referencias complejas
            peso_enc = 0.6   # Valor fijo para evitar referencias complejas
            
            # Fórmula simple: promedio ponderado
            formula = f"=(B{row}*{peso_hist} + C{row}*{peso_enc})"
            ws.cell(row=row, column=4, value=formula)
            
            # Aplicar estilos
            for col in range(1, 5):
                self.aplicar_estilo_celda(ws.cell(row=row, column=col))
        
        # Normalización de porcentajes
        last_row = len(all_parties) + 2
        ws.cell(row=last_row+1, column=3, value="Total (sin normalizar):")
        ws.cell(row=last_row+1, column=4, value=f"=SUM(D3:D{last_row})")
        
        ws.cell(row=last_row+2, column=3, value="Factor de normalización:")
        ws.cell(row=last_row+2, column=4, value=f"=100/D{last_row+1}")
        
        # Aplicar normalización
        for row in range(3, last_row+1):
            ws.cell(row=row, column=5, value="Predicción Normalizada")
            ws.cell(row=row, column=6, value=f"=D{row}*D${last_row+2}")
            if row == 3:
                self.aplicar_estilo_celda(ws.cell(row=row, column=5), is_header=True)
                self.aplicar_estilo_celda(ws.cell(row=row, column=6), is_header=True)
            else:
                self.aplicar_estilo_celda(ws.cell(row=row, column=5))
                self.aplicar_estilo_celda(ws.cell(row=row, column=6))
        
        # Verificación de segunda vuelta simplificada
        ws.cell(row=last_row+4, column=1, value="Verificación de Segunda Vuelta")
        ws.merge_cells(f'A{last_row+4}:D{last_row+4}')
        self.aplicar_estilo_celda(ws.cell(row=last_row+4, column=1), is_header=True)
        
        # Fórmulas simplificadas para verificar segunda vuelta
        ws.cell(row=last_row+5, column=1, value="Primer lugar:")
        ws.cell(row=last_row+5, column=2, value=f"=INDEX(A3:A{last_row}, MATCH(MAX(F3:F{last_row}), F3:F{last_row}, 0))")
        ws.cell(row=last_row+5, column=3, value="Votos:")
        ws.cell(row=last_row+5, column=4, value=f"=MAX(F3:F{last_row})")
        
        ws.cell(row=last_row+6, column=1, value="Segundo lugar:")
        ws.cell(row=last_row+6, column=2, value=f"=INDEX(A3:A{last_row}, MATCH(LARGE(F3:F{last_row}, 2), F3:F{last_row}, 0))")
        ws.cell(row=last_row+6, column=3, value="Votos:")
        ws.cell(row=last_row+6, column=4, value=f"=LARGE(F3:F{last_row}, 2)")
        
        # Fórmula simplificada para segunda vuelta
        ws.cell(row=last_row+7, column=1, value="¿Requiere segunda vuelta?")
        ws.cell(row=last_row+7, column=2, value=f"=IF(D{last_row+5}>50, \"NO\", IF(AND(D{last_row+5}>=40, D{last_row+5}-D{last_row+6}>=10), \"NO\", \"SÍ\"))")
        
        for row in range(last_row+5, last_row+8):
            for col in range(1, 5):
                if ws.cell(row=row, column=col).value:
                    self.aplicar_estilo_celda(ws.cell(row=row, column=col), is_subheader=(col == 1 or col == 3))
        
        # Gráfico de predicción
        chart = BarChart()
        chart.type = "col"
        chart.title = "Predicción de Votos 2025"
        chart.y_axis.title = "Porcentaje de Votos"
        chart.x_axis.title = "Partido Político"
        
        data = Reference(ws, min_col=6, min_row=2, max_row=last_row, max_col=6)
        categories = Reference(ws, min_col=1, min_row=3, max_row=last_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 15
        chart.width = 30
        
        ws.add_chart(chart, "H2")
        
        # Ajustar anchos de columna
        for col_idx in range(1, 7):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
    
    def crear_hoja_escanos(self):
        """Crea la hoja con la distribución de escaños usando método D'Hondt con columnas auxiliares, para Senadores y Diputados, máximo 20 divisores."""
        ws = self.ws_escanos
        self._crear_bloque_dhondt(ws, 'Senadores', self.total_senadores, start_row=1)
        # Espacio entre bloques
        self._crear_bloque_dhondt(ws, 'Diputados', self.total_diputados, start_row=60)

    def _crear_bloque_dhondt(self, ws, tipo, total_escanos, start_row=1):
        # Generar tantos cocientes como escaños a repartir
        max_div = total_escanos
        ws.cell(row=start_row, column=1, value=f"Distribución de {tipo} (Total: {total_escanos})")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        self.aplicar_estilo_celda(ws.cell(row=start_row, column=1), is_header=True)
        
        # Ajustar filas para evitar celdas combinadas
        tabla_inicio = start_row + 2
        coc_row_start = tabla_inicio + 1
        coc_col_start = 4

        # Obtener partidos y votos normalizados de la predicción
        pred_ws = self.ws_prediccion
        partidos = []
        votos = []
        for row in range(3, pred_ws.max_row):
            partido = pred_ws.cell(row=row, column=1).value
            voto = pred_ws.cell(row=row, column=6).value
            if isinstance(voto, str) and voto.startswith('='):
                votos.append(f"'Predicción 2025'!F{row}")
            else:
                votos.append(voto)
            partidos.append(partido)
        
        # Filtrar partidos que superan el umbral (configurable)
        umbral = self.umbral_minimo * 100
        partidos_validos = []
        votos_validos = []
        filas_validas = []
        for i, v in enumerate(votos):
            if isinstance(v, str) and v.startswith("'"):
                partidos_validos.append(partidos[i])
                votos_validos.append(v)
                filas_validas.append(i)
            else:
                try:
                    if float(v) >= umbral:
                        partidos_validos.append(partidos[i])
                        votos_validos.append(v)
                        filas_validas.append(i)
                except:
                    pass
        n = len(partidos_validos)

        # Tabla principal: partidos y votos válidos
        ws.cell(row=tabla_inicio, column=1, value="Partido")
        ws.cell(row=tabla_inicio, column=2, value="Votos (%)")
        self.aplicar_estilo_celda(ws.cell(row=tabla_inicio, column=1), is_header=True)
        self.aplicar_estilo_celda(ws.cell(row=tabla_inicio, column=2), is_header=True)
        for idx, (partido, voto, fila_pred) in enumerate(zip(partidos_validos, votos_validos, filas_validas)):
            row = tabla_inicio + 1 + idx
            ws.cell(row=row, column=1, value=f"='Predicción 2025'!A{fila_pred+3}")
            ws.cell(row=row, column=2, value=f"='Predicción 2025'!F{fila_pred+3}")
            self.aplicar_estilo_celda(ws.cell(row=row, column=1))
            self.aplicar_estilo_celda(ws.cell(row=row, column=2))

        # Tabla auxiliar de cocientes
        ws.cell(row=tabla_inicio, column=coc_col_start, value="Partido")
        ws.cell(row=tabla_inicio, column=coc_col_start+1, value="Cociente")
        ws.cell(row=tabla_inicio, column=coc_col_start+2, value="Divisor")
        ws.cell(row=tabla_inicio, column=coc_col_start+3, value="¿Escaño?")
        self.aplicar_estilo_celda(ws.cell(row=tabla_inicio, column=coc_col_start), is_header=True)
        self.aplicar_estilo_celda(ws.cell(row=tabla_inicio, column=coc_col_start+1), is_header=True)
        self.aplicar_estilo_celda(ws.cell(row=tabla_inicio, column=coc_col_start+2), is_header=True)
        self.aplicar_estilo_celda(ws.cell(row=tabla_inicio, column=coc_col_start+3), is_header=True)

        cocientes_refs = []
        filas_cocientes_por_partido = [[] for _ in range(n)]
        fila_coc = coc_row_start
        for idx in range(n):
            for div in range(1, max_div+1):
                # Buscar la siguiente fila libre (no combinada)
                while ws.cell(row=fila_coc, column=coc_col_start).coordinate in ws.merged_cells:
                    fila_coc += 1
                # En vez de referencia, escribe el nombre del partido directamente
                ws.cell(row=fila_coc, column=coc_col_start, value=partidos_validos[idx])
                # Si el voto es fórmula, usa fórmula; si es número, pon el valor calculado
                if isinstance(votos_validos[idx], str) and votos_validos[idx].startswith("'"):
                    voto_ref = votos_validos[idx]
                    ws.cell(row=fila_coc, column=coc_col_start+1, value=f"=IF({voto_ref}>0, {voto_ref}/{div}, 0)")
                else:
                    voto_val = votos_validos[idx]
                    ws.cell(row=fila_coc, column=coc_col_start+1, value=float(voto_val) / div if float(voto_val) > 0 else 0)
                ws.cell(row=fila_coc, column=coc_col_start+2, value=div)
                cocientes_refs.append(f"{get_column_letter(coc_col_start+1)}{fila_coc}")
                filas_cocientes_por_partido[idx].append(fila_coc)
                self.aplicar_estilo_celda(ws.cell(row=fila_coc, column=coc_col_start))
                self.aplicar_estilo_celda(ws.cell(row=fila_coc, column=coc_col_start+1))
                self.aplicar_estilo_celda(ws.cell(row=fila_coc, column=coc_col_start+2))
                fila_coc += 1
        # Columna auxiliar: 1 si este cociente está entre los N mayores
        cocientes_range = f"{get_column_letter(coc_col_start+1)}{coc_row_start}:{get_column_letter(coc_col_start+1)}{fila_coc-1}"
        fila_coc2 = coc_row_start
        for idx in range(n):
            for div in range(1, max_div+1):
                while ws.cell(row=fila_coc2, column=coc_col_start).coordinate in ws.merged_cells:
                    fila_coc2 += 1
                coc_ref = f"{get_column_letter(coc_col_start+1)}{fila_coc2}"
                ws.cell(row=fila_coc2, column=coc_col_start+3, value=f"=IFERROR(IF({coc_ref}>=LARGE({cocientes_range},{total_escanos}),1,0),0)")
                self.aplicar_estilo_celda(ws.cell(row=fila_coc2, column=coc_col_start+3))
                fila_coc2 += 1

        # Asignación de escaños: suma explícita de celdas
        result_row = fila_coc + 2
        ws.cell(row=result_row, column=1, value="Partido")
        ws.cell(row=result_row, column=2, value="Escaños")
        self.aplicar_estilo_celda(ws.cell(row=result_row, column=1), is_header=True)
        self.aplicar_estilo_celda(ws.cell(row=result_row, column=2), is_header=True)
        for idx in range(n):
            row = result_row + 1 + idx
            ws.cell(row=row, column=1, value=f"=A{tabla_inicio+1+idx}")
            # Suma explícita de celdas
            celdas = [f"{get_column_letter(coc_col_start+3)}{fila}" for fila in filas_cocientes_por_partido[idx]]
            ws.cell(row=row, column=2, value=f"=SUM({','.join(celdas)})")
            self.aplicar_estilo_celda(ws.cell(row=row, column=1))
            self.aplicar_estilo_celda(ws.cell(row=row, column=2))
        
        # Gráfico de distribución
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Distribución de {tipo}"
        chart.y_axis.title = "Número de Escaños"
        chart.x_axis.title = "Partido Político"
        data = Reference(ws, min_col=2, min_row=result_row+1, max_row=result_row+n)
        categories = Reference(ws, min_col=1, min_row=result_row+1, max_row=result_row+n)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.height = 15
        chart.width = 25
        ws.add_chart(chart, f"H{result_row}")
        
        # Nota explicativa
        ws.cell(row=result_row+n+2, column=1, value=f"* El reparto de escaños de {tipo} se calcula con la tabla auxiliar de cocientes a la derecha.")
        ws.merge_cells(start_row=result_row+n+2, start_column=1, end_row=result_row+n+2, end_column=6)
        self.aplicar_estilo_celda(ws.cell(row=result_row+n+2, column=1), is_subheader=True)
    
    def crear_hoja_segunda_vuelta(self):
        """Crea la hoja para simular segunda vuelta electoral simplificada"""
        ws = self.ws_segunda_vuelta
        
        # Título
        ws['A1'] = "Simulación de Segunda Vuelta Electoral"
        ws.merge_cells('A1:D1')
        self.aplicar_estilo_celda(ws['A1'], is_header=True)
        
        # Referencias a la hoja de predicción
        pred_ws = self.ws_prediccion
        # Buscar la última fila con partido en la hoja de predicción
        partidos = []
        row = 3
        while True:
            value = pred_ws.cell(row=row, column=1).value
            if value is None or value == "":
                break
            partidos.append(value)
            row += 1
        last_row = row - 1
        votos_col = 'F'
        partidos_col = 'A'
        
        # Verificar si se requiere segunda vuelta
        ws['A3'] = "¿Requiere segunda vuelta?"
        ws['B3'] = f'=IF({votos_col}3>50, "NO", IF(AND({votos_col}3>=40, {votos_col}3-{votos_col}4>=10), "NO", "SÍ"))'
        self.aplicar_estilo_celda(ws['A3'], is_subheader=True)
        self.aplicar_estilo_celda(ws['B3'])
        
        # Candidatos a segunda vuelta
        ws['A5'] = "Candidatos a Segunda Vuelta"
        ws.merge_cells('A5:D5')
        self.aplicar_estilo_celda(ws['A5'], is_header=True)
        
        ws['A6'] = "Primer lugar:"
        ws['B6'] = f'=INDEX({partidos_col}3:{partidos_col}{last_row}, MATCH(LARGE({votos_col}3:{votos_col}{last_row}, 1), {votos_col}3:{votos_col}{last_row}, 0))'
        ws['C6'] = "Votos:"
        ws['D6'] = f'=LARGE({votos_col}3:{votos_col}{last_row}, 1)'
        ws['A7'] = "Segundo lugar:"
        ws['B7'] = f'=INDEX({partidos_col}3:{partidos_col}{last_row}, MATCH(LARGE({votos_col}3:{votos_col}{last_row}, 2), {votos_col}3:{votos_col}{last_row}, 0))'
        ws['C7'] = "Votos:"
        ws['D7'] = f'=LARGE({votos_col}3:{votos_col}{last_row}, 2)'
        for col in range(1, 5):
            self.aplicar_estilo_celda(ws.cell(row=6, column=col), is_subheader=(col == 1 or col == 3))
            self.aplicar_estilo_celda(ws.cell(row=7, column=col), is_subheader=(col == 1 or col == 3))
        
        # Simulación de redistribución de votos simplificada
        ws['A9'] = "Simulación de Redistribución de Votos"
        ws.merge_cells('A9:D9')
        self.aplicar_estilo_celda(ws['A9'], is_header=True)
        
        ws['A10'] = "Partido"
        ws['B10'] = "Votos Primera Vuelta (%)"
        ws['C10'] = "Porcentaje redistribución"
        ws['D10'] = "Votos Segunda Vuelta (%)"
        for col in range(1, 5):
            self.aplicar_estilo_celda(ws.cell(row=10, column=col), is_header=True)
        
        # Primer candidato
        ws['A11'] = '=B6'
        ws['B11'] = '=D6'
        ws['C11'] = '70%'
        ws['D11'] = '=B11 + (100-B11-B12)*0.7'
        # Segundo candidato
        ws['A12'] = '=B7'
        ws['B12'] = '=D7'
        ws['C12'] = '30%'
        ws['D12'] = '=B12 + (100-B11-B12)*0.3'
        for row in range(11, 13):
            for col in range(1, 5):
                self.aplicar_estilo_celda(ws.cell(row=row, column=col))
        
        # Resultado final
        ws['A14'] = "Resultado Final"
        ws.merge_cells('A14:D14')
        self.aplicar_estilo_celda(ws['A14'], is_header=True)
        ws['A15'] = "Ganador:"
        ws['B15'] = '=IF(D11>D12, A11, A12)'
        ws['C15'] = "Porcentaje:"
        ws['D15'] = '=IF(D11>D12, D11, D12)'
        for col in range(1, 5):
            self.aplicar_estilo_celda(ws.cell(row=15, column=col), is_subheader=(col == 1 or col == 3))
        
        # Gráfico de resultados
        chart = BarChart()
        chart.type = "col"
        chart.title = "Resultados de Segunda Vuelta"
        chart.y_axis.title = "Porcentaje de Votos"
        chart.x_axis.title = "Candidato"
        data = Reference(ws, min_col=4, min_row=10, max_row=12)
        categories = Reference(ws, min_col=1, min_row=11, max_row=12)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 15
        chart.width = 20
        ws.add_chart(chart, "F3")
        
        # Ajustar anchos de columna
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
    
    def guardar_excel(self, filename="Prediccion_Electoral_Bolivia_2025.xlsx"):
        """Guarda el archivo Excel con todas las hojas"""
        self.crear_hoja_datos()
        self.crear_hoja_prediccion()
        self.crear_hoja_escanos()
        self.crear_hoja_segunda_vuelta()
        self.wb.save(filename)
        print(f"Archivo Excel generado exitosamente: {filename}")
        print("El archivo contiene fórmulas interactivas que se pueden modificar.")

# Ejecutar el generador
if __name__ == "__main__":
    generador = ExcelElectoralModel()
    generador.guardar_excel()